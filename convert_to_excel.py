import result
import openpyxl

courses = result.handbook["allCourses"]["courseDetails"]

# Extract fieldnames from the first course (assuming all courses have the same keys)
fieldnames = list(courses[0].keys())

# Define the name of the Excel file
excel_filename = 'all_courses.xlsx'

# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write header row
for col_idx, fieldname in enumerate(fieldnames, start=1):
    sheet.cell(row=1, column=col_idx, value=str(fieldname))

# Write data rows
for row_idx, course in enumerate(courses, start=2):
    for col_idx, fieldname in enumerate(fieldnames, start=1):
        value = course[fieldname]
        sheet.cell(row=row_idx, column=col_idx, value=str(value))

# Save workbook to Excel file
workbook.save(excel_filename)

print(f'Excel file "{excel_filename}" has been created successfully.')